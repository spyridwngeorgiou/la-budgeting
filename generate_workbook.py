"""
LA Budgeting — Excel Workbook Generator
Produces TWO files, each with 4 tabs:
  • LA_Budgeting_GR.xlsx  — fully in Greek
  • LA_Budgeting_EN.xlsx  — fully in English

Tabs:
  1. Control Center  — cockpit dashboard (read-only, 4-column list style)
  2. Expenses        — project expenses with Excel Tables (auto-expand)
  3. Accounts        — bank / cash / gold balances
  4. Info            — user guide in the file's language
"""
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo

# ─── COLOUR PALETTE — Clean Corporate (8-digit ARGB) ─────────────────────────
DARK_BLUE    = "FF1A2F4A"   # deep navy  — titles & primary headers
MED_BLUE     = "FF1868A8"   # ocean blue — sub-headers & accents
LIGHT_BLUE   = "FFE3EFF8"   # pale ice   — column headers & totals rows
YELLOW       = "FFFDF8ED"   # warm ivory — input / editable cells
LIGHT_GREY   = "FFF6F8FA"   # near-white — body row backgrounds
GREEN_BG     = "FFE8F8F1"   # mint       — Paid status background
GREEN_FG     = "FF0B6640"   # emerald    — Paid status text
RED_BG       = "FFFCEFEF"   # blush      — warning backgrounds
RED_FG       = "FF991B1B"   # crimson    — warning text
WHITE        = "FFFFFFFF"
BLACK        = "FF000000"
DARK_GREY    = "FF374151"   # slate      — secondary text
ORANGE_LIGHT = "FFFDF0D8"   # warm cream — Lazaraki section tint
PURPLE_LIGHT = "FFF0EEFF"   # soft lavender — Agiou K section tint
EUR = '\u20ac#,##0'
PCT = '0%'

# ─── LANGUAGE ────────────────────────────────────────────────────────────────
LANG = "gr"          # set per build; "gr" or "en"
def t(gr, en):
    """Return the Greek or English string for the current build language."""
    return en if LANG == "en" else gr

# ─── STYLE HELPERS ───────────────────────────────────────────────────────────
def F(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def Fnt(size=11, bold=False, color=None, italic=False):
    fg = color[2:] if color and color.startswith("FF") and len(color) == 8 else color
    return Font(name="Calibri", size=size, bold=bold,
                color=fg if fg else "000000", italic=italic)

def Aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def Brd(style="thin", color="FFCCD5DF"):
    s = Side(border_style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def col_hdr(ws, row, col, text, bg=LIGHT_BLUE, fg=DARK_GREY, size=10):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = F(bg); c.font = Fnt(size=size, bold=True, color=fg)
    c.alignment = Aln(h="center"); c.border = Brd()

def inp(ws, row, col, value, fmt=None, align_h="right"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = F(YELLOW); c.font = Fnt(); c.alignment = Aln(h=align_h)
    c.border = Brd()
    if fmt: c.number_format = fmt
    return c

def calc(ws, row, col, formula, fmt=None, bg=LIGHT_GREY, bold=False, size=11, fg=None):
    c = ws.cell(row=row, column=col, value=formula)
    c.fill = F(bg); c.font = Fnt(size=size, bold=bold, color=fg if fg else DARK_GREY)
    c.alignment = Aln(h="right"); c.border = Brd()
    if fmt: c.number_format = fmt
    return c

def lbl(ws, row, col, text, bg=LIGHT_GREY, bold=False, size=11, h="left"):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = F(bg); c.font = Fnt(size=size, bold=bold)
    c.alignment = Aln(h=h); c.border = Brd()
    return c

# ─── LAYOUT HELPERS ──────────────────────────────────────────────────────────
def mrow(ws, row, label, value, note="",
         label_bg=WHITE, value_bg=WHITE, note_bg=WHITE,
         label_bold=False, value_bold=False, value_size=11,
         value_color=None, value_fmt=None,
         accent=DARK_BLUE, height=22):
    ws.row_dimensions[row].height = height
    a = ws.cell(row=row, column=1)
    a.fill = F(accent); a.border = Brd(color=accent)
    b = ws.cell(row=row, column=2, value=label)
    b.font = Fnt(size=11, bold=label_bold, color="FF374151")
    b.alignment = Aln(h="left", v="center", wrap=True); b.fill = F(label_bg)
    b.border = Brd(color="FFCCD5DF")
    c = ws.cell(row=row, column=3, value=value)
    fg = value_color if value_color else "FF374151"
    c.font = Fnt(size=value_size, bold=value_bold, color=fg)
    c.alignment = Aln(h="right", v="center"); c.fill = F(value_bg)
    c.border = Brd(color="FFCCD5DF")
    if value_fmt: c.number_format = value_fmt
    d = ws.cell(row=row, column=4, value=note)
    d.font = Fnt(size=10, italic=True, color="FF6B7280")
    d.alignment = Aln(h="left", v="center", wrap=True); d.fill = F(note_bg)
    d.border = Brd(color="FFCCD5DF")
    return c

def sec_hdr(ws, row, text, bg=DARK_BLUE, height=24):
    ws.row_dimensions[row].height = height
    a = ws.cell(row=row, column=1)
    a.fill = F(bg)
    ws.merge_cells(f"B{row}:D{row}")
    c = ws.cell(row=row, column=2, value=f"  {text}")
    c.fill = F(LIGHT_GREY)
    c.font = Fnt(size=10, bold=True, color=DARK_BLUE)
    c.alignment = Aln(h="left", v="center")
    c.border = Border(
        top=Side(border_style="thin",    color="FFCCD5DF"),
        bottom=Side(border_style="medium", color=bg),
        left=Side(border_style="thin",    color="FFCCD5DF"),
        right=Side(border_style="thin",   color="FFCCD5DF"))

def spacer(ws, row, height=8, bg=WHITE):
    ws.row_dimensions[row].height = height
    for ci in range(1, 5): ws.cell(row=row, column=ci).fill = F(bg)

def exp_metric(ws, row, label, formula, note, bg, fg=None, height=24):
    ws.row_dimensions[row].height = height
    c = ws.cell(row=row, column=1, value=label)
    c.font = Fnt(size=11, bold=True, color=fg if fg else "FF374151")
    c.fill = F(bg); c.border = Brd(color="FFCCD5DF")
    c.alignment = Aln(h="left", v="center")
    for ci in [2, 3]:
        ws.cell(row=row, column=ci).fill = F(bg)
        ws.cell(row=row, column=ci).border = Brd(color="FFCCD5DF")
    val = ws.cell(row=row, column=4, value=formula)
    val.font = Fnt(size=14, bold=True, color=fg if fg else "FF374151")
    val.fill = F(bg); val.border = Brd(color="FFCCD5DF")
    val.number_format = EUR; val.alignment = Aln(h="right", v="center")
    nt = ws.cell(row=row, column=5, value=note)
    nt.font = Fnt(size=10, italic=True, color="FF6B7280")
    nt.fill = F(bg); nt.border = Brd(color="FFCCD5DF")
    nt.alignment = Aln(h="left", v="center")
    ws.cell(row=row, column=6).fill = F(bg)
    ws.cell(row=row, column=6).border = Brd(color="FFCCD5DF")

def exp_data_row(ws, row, project, item, status, source, amt, notes):
    ws.row_dimensions[row].height = 18
    proj_bg = LIGHT_BLUE if "Laz" in project else PURPLE_LIGHT
    proj_fg = "FF1A2F4A" if "Laz" in project else "FF3E1A7C"
    status_bg = GREEN_BG if status in ("\u03a0\u03bb\u03b7\u03c1\u03c9\u03bc\u03ad\u03bd\u03bf", "Paid") else YELLOW
    c = ws.cell(row=row, column=1, value=project)
    c.fill = F(proj_bg); c.alignment = Aln(h="center", v="center"); c.border = Brd()
    c.font = Fnt(size=9, bold=True, color=proj_fg)
    ws.cell(row=row, column=2, value=item).fill = F(WHITE)
    ws.cell(row=row, column=2).border = Brd()
    ws.cell(row=row, column=2).alignment = Aln(h="left", v="center")
    c = ws.cell(row=row, column=3, value=status)
    c.fill = F(status_bg); c.alignment = Aln(h="center"); c.border = Brd()
    ws.cell(row=row, column=4, value=source).fill = F(WHITE)
    ws.cell(row=row, column=4).alignment = Aln(h="center"); ws.cell(row=row, column=4).border = Brd()
    c = ws.cell(row=row, column=5, value=amt)
    c.number_format = EUR; c.alignment = Aln(h="right"); c.border = Brd(); c.fill = F(WHITE)
    ws.cell(row=row, column=6, value=notes).fill = F(WHITE)
    ws.cell(row=row, column=6).alignment = Aln(wrap=True); ws.cell(row=row, column=6).border = Brd()

def exp_empty_rows(ws, start, end):
    for r in range(start, end + 1):
        ws.row_dimensions[r].height = 18
        ws.cell(row=r, column=1).fill = F(YELLOW)
        ws.cell(row=r, column=1).alignment = Aln(h="center")
        ws.cell(row=r, column=1).border = Brd(color="FFD0D0D0")
        ws.cell(row=r, column=2).fill = F(WHITE); ws.cell(row=r, column=2).border = Brd(color="FFD0D0D0")
        ws.cell(row=r, column=3).fill = F(YELLOW)
        ws.cell(row=r, column=3).alignment = Aln(h="center")
        ws.cell(row=r, column=3).border = Brd(color="FFD0D0D0")
        ws.cell(row=r, column=4).fill = F(YELLOW)
        ws.cell(row=r, column=4).alignment = Aln(h="center")
        ws.cell(row=r, column=4).border = Brd(color="FFD0D0D0")
        ws.cell(row=r, column=5).fill = F(YELLOW)
        ws.cell(row=r, column=5).alignment = Aln(h="right")
        ws.cell(row=r, column=5).number_format = EUR
        ws.cell(row=r, column=5).border = Brd(color="FFD0D0D0")
        ws.cell(row=r, column=6).fill = F(WHITE); ws.cell(row=r, column=6).border = Brd(color="FFD0D0D0")

def _info_tab(wb, title, tab_title, sections):
    """Build a styled info/guide tab."""
    ws = wb.create_sheet(tab_title)
    ws.sheet_view.showGridLines = False
    B_W, C_W = 60, 32
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = B_W
    ws.column_dimensions["C"].width = C_W
    ws.column_dimensions["D"].width = 4

    def est_lines(s, width):
        """Estimate how many wrapped lines `s` needs in a column `width` chars wide."""
        if not s:
            return 1
        cpl = max(8, int(width * 1.05))
        n, cur = 1, 0
        for word in str(s).split(" "):
            w = len(word) + 1
            if cur + w > cpl and cur > 0:
                n += 1
                cur = w
            else:
                cur += w
        return n

    ACCENT = "FF1A2F4A"
    ws.row_dimensions[1].height = 40
    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1, value=f"  {title}")
    c.fill = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
    c.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[2].height = 6
    for ci in range(1, 5):
        ws.cell(row=2, column=ci).fill = PatternFill(start_color="FFF6F8FA",
                                                     end_color="FFF6F8FA", fill_type="solid")

    row = 3
    thin = Side(border_style="thin", color="FFCCD5DF")

    for sec_title, sec_color, items in sections:
        ws.row_dimensions[row].height = 26
        ws.cell(row=row, column=1).fill = PatternFill(start_color=sec_color,
                                                      end_color=sec_color, fill_type="solid")
        ws.merge_cells(f"B{row}:D{row}")
        c = ws.cell(row=row, column=2, value=f"  {sec_title}")
        c.fill = PatternFill(start_color="FFF6F8FA", end_color="FFF6F8FA", fill_type="solid")
        c.font = Font(name="Calibri", size=11, bold=True, color="1A2F4A")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = Border(top=thin, bottom=Side(border_style="medium", color=sec_color),
                          left=thin, right=thin)
        row += 1

        for (rtype, text, note) in items:
            if rtype == "sp":
                ws.row_dimensions[row].height = int(text) if text else 8
                for ci in range(1, 5):
                    ws.cell(row=row, column=ci).fill = PatternFill(
                        start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
                row += 1
                continue

            a = ws.cell(row=row, column=1)

            if rtype == "h":
                bg = "FFE3EFF8"; fg = "1A2F4A"; bold = True; size = 11
                a.fill = PatternFill(start_color="FF1868A8", end_color="FF1868A8", fill_type="solid")
            elif rtype == "s":
                bg = "FFFFFFFF"; fg = "374151"; bold = False; size = 11
                a.fill = PatternFill(start_color="FF1A2F4A", end_color="FF1A2F4A", fill_type="solid")
            elif rtype == "tip":
                bg = "FFFDF8ED"; fg = "7A3800"; bold = False; size = 10
                a.fill = PatternFill(start_color="FFFF9800", end_color="FFFF9800", fill_type="solid")
            elif rtype == "clr":
                bg, fg, bold, size = text[0], text[1], False, 10
                a.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                a.border = Border(top=thin, bottom=thin, left=thin, right=thin)
                ws.merge_cells(f"B{row}:D{row}")
                bcell = ws.cell(row=row, column=2, value=f"  {note}")
                bcell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                fgc = fg[2:] if fg.startswith("FF") and len(fg) == 8 else fg
                bcell.font = Font(name="Calibri", size=10, bold=True, color=fgc)
                bcell.alignment = Alignment(horizontal="left", vertical="center")
                bcell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
                ws.cell(row=row, column=4).fill = PatternFill(
                    start_color=bg, end_color=bg, fill_type="solid")
                ws.row_dimensions[row].height = 22
                row += 1
                continue
            else:
                bg = "FFFFFFFF"; fg = "374151"; bold = False; size = 11
                a.fill = PatternFill(start_color="FFF6F8FA", end_color="FFF6F8FA", fill_type="solid")

            a.border = Border(top=thin, bottom=thin, left=thin, right=thin)

            has_note = bool(note)
            if not has_note:
                ws.merge_cells(f"B{row}:C{row}")
            main_w = (B_W + C_W) if not has_note else B_W

            bcell = ws.cell(row=row, column=2, value=text)
            fgc = fg[2:] if fg.startswith("FF") and len(fg) == 8 else fg
            bcell.font = Font(name="Calibri", size=size, bold=bold, color=fgc)
            bcell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            bcell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

            ncell = None
            if has_note:
                ncell = ws.cell(row=row, column=3, value=note)
                ncell.font = Font(name="Calibri", size=10, italic=True, color="6B7280")
                ncell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                ncell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

            dcell = ws.cell(row=row, column=4)
            dcell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            dcell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

            # dynamic row height so wrapped guide text is never clipped
            lines = max(est_lines(text, main_w), est_lines(note, C_W))
            per = 15 if size >= 11 else 14
            ws.row_dimensions[row].height = max(20, lines * per + 7)
            vtop = "top" if lines > 1 else "center"
            bcell.alignment = Alignment(horizontal="left", vertical=vtop, wrap_text=True)
            if ncell is not None:
                ncell.alignment = Alignment(horizontal="left", vertical=vtop, wrap_text=True)
            row += 1

        ws.row_dimensions[row].height = 10
        for ci in range(1, 5):
            ws.cell(row=row, column=ci).fill = PatternFill(
                start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
        row += 1

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# BUILD ONE WORKBOOK (language driven by the global LANG via t())
# ══════════════════════════════════════════════════════════════════════════════
def build_workbook():
    # localized status / source values (also used inside formulas & dropdowns)
    S_PAID   = t("\u03a0\u03bb\u03b7\u03c1\u03c9\u03bc\u03ad\u03bd\u03bf", "Paid")
    S_UPCOM  = t("\u0395\u03ba\u03ba\u03c1\u03b5\u03bc\u03b5\u03af", "Upcoming")
    S_HOLD   = t("\u03a3\u03b5 \u03b1\u03bd\u03b1\u03bc\u03bf\u03bd\u03ae", "On Hold")
    SRC_PIR  = t("\u03a4\u03c1\u03ac\u03c0\u03b5\u03b6\u03b1 \u03a0\u03b5\u03b9\u03c1\u03b1\u03b9\u03ce\u03c2", "Piraeus Bank")
    SRC_CASH = t("\u039c\u03b5\u03c4\u03c1\u03b7\u03c4\u03ac", "Cash")
    SRC_TBD  = t("\u03a0\u03c1\u03bf\u03c2 \u03ba\u03b1\u03b8\u03bf\u03c1\u03b9\u03c3\u03bc\u03cc", "TBD")
    SRC_GOLD = t("\u03a7\u03c1\u03c5\u03c3\u03cc\u03c2", "Gold")
    A_PEND   = t("\u0395\u03ba\u03ba\u03c1\u03b5\u03bc\u03b5\u03af", "Pending")
    A_RECV   = t("\u0395\u03bb\u03ae\u03c6\u03b8\u03b7", "Received")
    A_CANC   = t("\u0391\u03ba\u03c5\u03c1\u03ce\u03b8\u03b7\u03ba\u03b5", "Cancelled")

    wb = Workbook()
    ws_cc  = wb.active;          ws_cc.title = "Control Center"
    ws_exp = wb.create_sheet("Expenses")
    ws_acc = wb.create_sheet("Accounts")

    # ══════════════════════════════════════════════════════════════════════════
    # TAB: CONTROL CENTER
    # ══════════════════════════════════════════════════════════════════════════
    ws = ws_cc
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 40

    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1,
                value="  " + t("\u03a0\u03a1\u039f\u03ab\u03a0\u039f\u039b\u039f\u0393\u0399\u03a3\u039c\u039f\u03a3 LA  \u2014  \u039a\u0395\u039d\u03a4\u03a1\u039f \u0395\u039b\u0395\u0393\u03a7\u039f\u03a5",
                          "LA BUDGETING  \u2014  CONTROL CENTER"))
    c.fill = F(DARK_BLUE); c.font = Fnt(size=15, bold=True, color=WHITE)
    c.alignment = Aln(h="left", v="center")

    ws.row_dimensions[2].height = 24
    ws.cell(row=2, column=1).fill = F("FF059669")
    b = ws.cell(row=2, column=2, value=t("\u039a\u03b1\u03c4\u03ac\u03c3\u03c4\u03b1\u03c3\u03b7 \u03a3\u03c5\u03c3\u03c4\u03ae\u03bc\u03b1\u03c4\u03bf\u03c2:", "System Status:"))
    b.font = Fnt(size=11, bold=True); b.alignment = Aln(h="left", v="center")
    b.fill = F("FFF0FAF8"); b.border = Brd(color="FFCCD5DF")
    ST_OK  = t("\u039b\u0395\u0399\u03a4\u039f\u03a5\u03a1\u0393\u0399\u039a\u039f", "OPERATIONAL")
    ST_ACT = t("\u0391\u03a0\u0391\u0399\u03a4\u0395\u0399\u03a4\u0391\u0399 \u0395\u039d\u0395\u03a1\u0393\u0395\u0399\u0391", "ACTION REQUIRED")
    c = ws.cell(row=2, column=3,
                value='=IF(Accounts!B10-Expenses!D5>=0,'
                      f'"{ST_OK}","{ST_ACT}")')
    c.font = Fnt(size=11, bold=True, color="FF059669")
    c.alignment = Aln(h="right", v="center"); c.fill = F("FFF0FAF8")
    c.border = Brd(color="FFCCD5DF")
    d = ws.cell(row=2, column=4,
                value='="' + t("\u03a4\u03b5\u03bb\u03b5\u03c5\u03c4\u03b1\u03af\u03b1 \u0395\u03bd\u03b7\u03bc\u03ad\u03c1\u03c9\u03c3\u03b7:  ", "Last Updated:  ") + '" & TEXT(TODAY(),"DD-MMM-YYYY")')
    d.font = Fnt(size=10, italic=True, color="FF6B7280")
    d.alignment = Aln(h="left", v="center"); d.fill = F("FFF0FAF8")
    d.border = Brd(color="FFCCD5DF")

    spacer(ws, 3, 6, "FFF0FAF8")

    # ── AVAILABLE FUNDS ──
    sec_hdr(ws, 4, t("\u0394\u0399\u0391\u0398\u0395\u03a3\u0399\u039c\u0391 \u039a\u0395\u03a6\u0391\u039b\u0391\u0399\u0391", "AVAILABLE FUNDS"))
    spacer(ws, 5, 5, LIGHT_GREY)
    mrow(ws, 6, "\U0001f4b0  " + t("\u039c\u03b5\u03c4\u03c1\u03b7\u03c4\u03ac \u03c3\u03c4\u03bf \u03c7\u03ad\u03c1\u03b9", "Cash in hand"), "=Accounts!B7",
         t("\u0394\u03b9\u03b1\u03b8\u03ad\u03c3\u03b9\u03bc\u03b1 \u03bc\u03b5\u03c4\u03c1\u03b7\u03c4\u03ac", "Physical cash available"),
         value_fmt=EUR, value_bold=True, value_size=12, value_color=MED_BLUE,
         label_bg=LIGHT_GREY, value_bg=WHITE)
    mrow(ws, 7, "\U0001f3e6  " + t("\u03a4\u03c1\u03ac\u03c0\u03b5\u03b6\u03b5\u03c2 (Optima + \u03a0\u03b5\u03b9\u03c1\u03b1\u03b9\u03ce\u03c2)", "Banks (Optima + Piraeus)"), "=Accounts!B5+Accounts!B6",
         t("\u03a5\u03c0\u03cc\u03bb\u03bf\u03b9\u03c0\u03b1 \u03c4\u03c1\u03b1\u03c0\u03b5\u03b6\u03b9\u03ba\u03ce\u03bd \u03bb\u03bf\u03b3\u03b1\u03c1\u03b9\u03b1\u03c3\u03bc\u03ce\u03bd", "Bank account balances"),
         value_fmt=EUR, value_bold=True, value_size=12, value_color=MED_BLUE,
         label_bg=LIGHT_GREY, value_bg=WHITE)
    mrow(ws, 8, "\U0001f947  " + t("\u0391\u03c0\u03bf\u03b8\u03ad\u03bc\u03b1\u03c4\u03b1 \u03c7\u03c1\u03c5\u03c3\u03bf\u03cd", "Gold reserves"), "=Accounts!B8",
         t("\u0395\u03ba\u03c4\u03b9\u03bc\u03ce\u03bc\u03b5\u03bd\u03b7 \u03b1\u03be\u03af\u03b1 \u03c7\u03c1\u03c5\u03c3\u03bf\u03cd", "Estimated gold value"),
         value_fmt=EUR, value_bold=True, value_size=12, value_color=MED_BLUE,
         label_bg=LIGHT_GREY, value_bg=WHITE)
    mrow(ws, 9, "\u2726  " + t("\u03a3\u03a5\u039d\u039f\u039b\u039f \u0395\u03a0\u0399\u0392\u0395\u0392\u0391\u0399\u03a9\u039c\u0395\u039d\u03a9\u039d \u039a\u0395\u03a6\u0391\u039b\u0391\u0399\u03a9\u039d", "TOTAL CONFIRMED FUNDS"), "=Accounts!B10",
         t("Όλοι οι λογαριασμοί (μαζί με έξτρα)", "All accounts (incl. extra)"),
         value_fmt=EUR, value_bold=True, value_size=14, value_color=DARK_BLUE,
         label_bg=LIGHT_BLUE, value_bg=LIGHT_BLUE, note_bg=LIGHT_BLUE,
         label_bold=True, height=30)
    spacer(ws, 10, 5, LIGHT_GREY)
    mrow(ws, 11, "\u23f3  " + t("Δάνειο TEPIX", "TEPIX Loan"), "=Accounts!B14",
         '="' + t("\u039a\u03b1\u03c4\u03ac\u03c3\u03c4\u03b1\u03c3\u03b7: ", "Status: ") + '" & Accounts!C14',
         value_fmt=EUR, value_bold=True, value_size=12, value_color="FF965800",
         label_bg="FFFEF0D8", value_bg="FFFEF0D8", note_bg="FFFEF0D8")
    mrow(ws, 12, "\u23f3  " + t("\u0391\u03bd\u03b1\u03bc\u03b5\u03bd\u03cc\u03bc\u03b5\u03bd\u03b1 (\u03ac\u03bb\u03bb\u03b1 \u03b5\u03b9\u03c3\u03b5\u03c1\u03c7\u03cc\u03bc\u03b5\u03bd\u03b1)", "Expecting (other incoming)"), "=Accounts!B15",
         '="' + t("\u039a\u03b1\u03c4\u03ac\u03c3\u03c4\u03b1\u03c3\u03b7: ", "Status: ") + '" & Accounts!C15',
         value_fmt=EUR, value_color="FF6B7280",
         label_bg=LIGHT_GREY, value_bg=LIGHT_GREY)
    spacer(ws, 13, 6)

    # ── helper for dynamic (auto-hiding) breakdown rows, reused lower down ──
    white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
    white_font = Font(name="Calibri", color="FFFFFFFF")

    def _acc_line(row, name_ref, value_ref, note_formula, emoji, color):
        ws.row_dimensions[row].height = 20
        ws.cell(row=row, column=1).fill = F(WHITE)
        fc = ws.cell(row=row, column=6, value=f'=IF({name_ref}="","",{name_ref})')
        fc.font = Fnt(size=8, color="FFFFFFFF")
        b = ws.cell(row=row, column=2, value=f'=IF({name_ref}="","","{emoji}  "&{name_ref})')
        b.fill = F(LIGHT_GREY); b.font = Fnt(size=11, color="FF374151")
        b.alignment = Aln(h="left", v="center", wrap=True)
        vc = ws.cell(row=row, column=3, value=f'=IF({name_ref}="","",{value_ref})')
        vc.fill = F(WHITE); vc.font = Fnt(size=12, bold=True, color=color)
        vc.alignment = Aln(h="right", v="center"); vc.number_format = EUR
        d = ws.cell(row=row, column=4, value=note_formula)
        d.fill = F(WHITE); d.font = Fnt(size=10, italic=True, color="FF6B7280")
        d.alignment = Aln(h="left", v="center", wrap=True)
        ws.cell(row=row, column=5).fill = F(WHITE)
        ws.conditional_formatting.add(f"A{row}:F{row}",
            FormulaRule(formula=[f'$F${row}=""'], fill=white_fill, font=white_font))

    STATUS_W = t("\u039a\u03b1\u03c4\u03ac\u03c3\u03c4\u03b1\u03c3\u03b7", "Status")

    # ── OVERVIEW (headline totals — always correct, aggregates everything) ──
    sec_hdr(ws, 14, t("\u0393\u0395\u039d\u0399\u039a\u0397 \u0395\u0399\u039a\u039f\u039d\u0391", "OVERVIEW"))
    spacer(ws, 15, 5, LIGHT_GREY)
    mrow(ws, 16, "\U0001f4bc  " + t("\u03a3\u03cd\u03bd\u03bf\u03bb\u03bf \u0394\u03b9\u03b1\u03b8\u03ad\u03c3\u03b9\u03bc\u03c9\u03bd \u039a\u03b5\u03c6\u03b1\u03bb\u03b1\u03af\u03c9\u03bd", "Total Available Funds"), "=Accounts!B10",
         t("\u038c\u03bb\u03bf\u03b9 \u03bf\u03b9 \u03bb\u03bf\u03b3\u03b1\u03c1\u03b9\u03b1\u03c3\u03bc\u03bf\u03af (\u03bc\u03b1\u03b6\u03af \u03bc\u03b5 \u03ad\u03be\u03c4\u03c1\u03b1)", "All accounts (incl. extra)"),
         value_fmt=EUR, value_bold=True, value_size=12, value_color=MED_BLUE, label_bg=LIGHT_GREY, value_bg=WHITE)
    mrow(ws, 17, "\u23f3  " + t("\u03a3\u03cd\u03bd\u03bf\u03bb\u03bf \u0395\u03b9\u03c3\u03b5\u03c1\u03c7\u03bf\u03bc\u03ad\u03bd\u03c9\u03bd", "Total Incoming"), "=Accounts!B17",
         t("TEPIX + \u03b1\u03bd\u03b1\u03bc\u03b5\u03bd\u03cc\u03bc\u03b5\u03bd\u03b1 + \u03ad\u03be\u03c4\u03c1\u03b1", "TEPIX + expecting + extra"),
         value_fmt=EUR, value_color="FF6B7280", label_bg=LIGHT_GREY, value_bg=WHITE)
    mrow(ws, 18, "\u2705  " + t("Σύνολο Πληρωμένων", "Total Paid"), "=Expenses!D4",
         t("Πληρωμένα όλων των έργων", "Paid across all projects"),
         value_fmt=EUR, value_color=GREEN_FG, label_bg=LIGHT_GREY, value_bg=WHITE)
    mrow(ws, 19, "\U0001f9fe  " + t("Σύνολο Υποχρεώσεων", "Total Committed"), "=Expenses!D5",
         t("Εκκρεμή όλων των έργων", "Upcoming across all projects"),
         value_fmt=EUR, value_color="FF7A3800", label_bg=LIGHT_GREY, value_bg=WHITE)
    mrow(ws, 20, "\U0001f4ca  " + t("Γενικός Προϋπολογισμός", "Grand Total Budget"), "=Expenses!D6",
         t("Σύνολο όλων των έργων (πληρωμένα + εκκρεμή)", "All projects (paid + upcoming)"),
         value_fmt=EUR, value_bold=True, value_size=12, value_color=DARK_BLUE, label_bg=LIGHT_GREY, value_bg=WHITE)
    mrow(ws, 21, "\U0001f4b5  " + t("\u0395\u039b\u0395\u03a5\u0398\u0395\u03a1\u039f \u039c\u0395\u03a4\u03a1\u0397\u03a4\u039f", "FREE CASH"),
         "=Accounts!B10-Expenses!D5",
         t("\u0394\u03b9\u03b1\u03b8\u03ad\u03c3\u03b9\u03bc\u03b1 \u03bc\u03b5\u03c4\u03ac \u03b1\u03c0\u03cc \u03cc\u03bb\u03b5\u03c2 \u03c4\u03b9\u03c2 \u03c5\u03c0\u03bf\u03c7\u03c1\u03b5\u03ce\u03c3\u03b5\u03b9\u03c2", "Available after all commitments"),
         value_fmt=EUR, value_bold=True, value_size=16, value_color=GREEN_FG,
         label_bg=GREEN_BG, value_bg=GREEN_BG, note_bg=GREEN_BG, label_bold=True, height=34)
    spacer(ws, 22, 6)

    # ── labels reused by the parametric project blocks below ──
    TOT_BUDGET = t("\u03a3\u03c5\u03bd\u03bf\u03bb\u03b9\u03ba\u03cc\u03c2 \u03a0\u03c1\u03bf\u03cb\u03c0\u03bf\u03bb\u03bf\u03b3\u03b9\u03c3\u03bc\u03cc\u03c2", "Total Budget")
    GRAND_NOTE = t("\u0393\u03b5\u03bd\u03b9\u03ba\u03cc \u03c3\u03cd\u03bd\u03bf\u03bb\u03bf (\u03c0\u03bb\u03b7\u03c1\u03c9\u03bc\u03ad\u03bd\u03b1 + \u03b5\u03ba\u03ba\u03c1\u03b5\u03bc\u03ae)", "Grand total (paid + upcoming)")
    ALREADY   = t("\u0389\u03b4\u03b7 \u03a0\u03bb\u03b7\u03c1\u03c9\u03bc\u03ad\u03bd\u03b1", "Already Paid")
    PAID_NOTE = t("\u03a0\u03bb\u03b7\u03c1\u03c9\u03bc\u03ad\u03c2 \u03bc\u03ad\u03c7\u03c1\u03b9 \u03c3\u03ae\u03bc\u03b5\u03c1\u03b1", "Payments made to date")
    REMAIN    = t("\u03a5\u03c0\u03cc\u03bb\u03bf\u03b9\u03c0\u03b5\u03c2 \u03a5\u03c0\u03bf\u03c7\u03c1\u03b5\u03ce\u03c3\u03b5\u03b9\u03c2", "Remaining Commitments")
    REM_NOTE  = t("\u0395\u03ba\u03ba\u03c1\u03b5\u03bc\u03ae / \u03c0\u03c1\u03bf\u03c2 \u03c0\u03bb\u03b7\u03c1\u03c9\u03bc\u03ae", "Upcoming / still to pay")
    PROGRESS  = t("\u03a0\u03c1\u03cc\u03bf\u03b4\u03bf\u03c2", "Progress")
    PROG_NOTE = t("\u03a0\u03bf\u03c3\u03bf\u03c3\u03c4\u03cc \u03c0\u03c1\u03bf\u03cb\u03c0\u03bf\u03bb\u03bf\u03b3\u03b9\u03c3\u03bc\u03bf\u03cd \u03c0\u03bf\u03c5 \u03c0\u03bb\u03b7\u03c1\u03ce\u03b8\u03b7\u03ba\u03b5", "Percentage of total budget paid")

    # ── (NET POSITION & ALERTS removed — kept fully parametric) ──

    # ── (detailed breakdown removed — Control Center shows totals only) ──


    # ── Conditional formatting ──
    red_fill   = PatternFill(start_color=RED_BG,   end_color=RED_BG,   fill_type="solid")
    green_fill = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
    for addr, sz in [("C21", 16)]:
        ws.conditional_formatting.add(addr,
            CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=green_fill,
                       font=Font(name="Calibri", size=sz, bold=True, color="0B6640")))
        ws.conditional_formatting.add(addr,
            CellIsRule(operator="lessThan", formula=["0"], fill=red_fill,
                       font=Font(name="Calibri", size=sz, bold=True, color="991B1B")))
    teal_fill = PatternFill(start_color="FFD1FAE5", end_color="FFD1FAE5", fill_type="solid")
    orng_fill = PatternFill(start_color="FFFEF0D8", end_color="FFFEF0D8", fill_type="solid")
    ws.conditional_formatting.add("C2",
        FormulaRule(formula=[f'C2="{ST_OK}"'], fill=teal_fill,
                    font=Font(name="Calibri", size=11, bold=True, color="059669")))
    ws.conditional_formatting.add("C2",
        FormulaRule(formula=[f'C2="{ST_ACT}"'], fill=orng_fill,
                    font=Font(name="Calibri", size=11, bold=True, color="965800")))

    # ══════════════════════════════════════════════════════════════════════════
    # TAB: EXPENSES
    # ══════════════════════════════════════════════════════════════════════════
    ws = ws_exp
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 30
    ws.freeze_panes = "A13"

    ws.row_dimensions[1].height = 34
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value="  " + t("\u039f\u039b\u0395\u03a3 \u039f\u0399 \u0394\u0391\u03a0\u0391\u039d\u0395\u03a3 \u0395\u03a1\u0393\u03a9\u039d", "ALL PROJECT EXPENSES"))
    c.fill = F(DARK_BLUE); c.font = Fnt(size=14, bold=True, color=WHITE)
    c.alignment = Aln(h="left", v="center")

    ws.row_dimensions[2].height = 8
    for ci in range(1, 7): ws.cell(row=2, column=ci).fill = F(LIGHT_GREY)

    M_PAID   = "\u2705  " + t("\u03a3\u03cd\u03bd\u03bf\u03bb\u03bf \u03a0\u03bb\u03b7\u03c1\u03c9\u03bc\u03ad\u03bd\u03c9\u03bd", "Total Paid")
    M_UPCOM  = "\u23f0  " + t("\u03a3\u03cd\u03bd\u03bf\u03bb\u03bf \u0395\u03ba\u03ba\u03c1\u03b5\u03bc\u03ce\u03bd", "Total Upcoming")
    M_GRAND  = "\U0001f4ca  " + t("\u0393\u03b5\u03bd\u03b9\u03ba\u03cc\u03c2 \u03a0\u03c1\u03bf\u03cb\u03c0\u03bf\u03bb\u03bf\u03b3\u03b9\u03c3\u03bc\u03cc\u03c2", "Grand Total Budget")
    M_PAID_N = t("\u03a0\u03bb\u03b7\u03c1\u03c9\u03bc\u03ad\u03c2 \u03bc\u03ad\u03c7\u03c1\u03b9 \u03c3\u03ae\u03bc\u03b5\u03c1\u03b1 (\u03cc\u03bb\u03b1 \u03c4\u03b1 \u03ad\u03c1\u03b3\u03b1)", "Payments to date (all projects)")
    M_UP_N   = t("\u0395\u03ba\u03ba\u03c1\u03b5\u03bc\u03ae \u03cc\u03bb\u03c9\u03bd \u03c4\u03c9\u03bd \u03ad\u03c1\u03b3\u03c9\u03bd", "Upcoming across all projects")

    # Table column names (localized) — used both as headers AND inside SUM formulas
    # structured references, so the Greek file resolves correctly (no #REF!).
    H_PROJECT = t("\u0388\u03c1\u03b3\u03bf", "Project")
    H_ITEM    = t("\u0395\u03af\u03b4\u03bf\u03c2 / \u03a0\u03b5\u03c1\u03b9\u03b3\u03c1\u03b1\u03c6\u03ae", "Item / Description")
    H_STATUS  = t("\u039a\u03b1\u03c4\u03ac\u03c3\u03c4\u03b1\u03c3\u03b7", "Status")
    H_SOURCE  = t("\u03a0\u03b7\u03b3\u03ae", "Source")
    H_AMOUNT  = t("\u03a0\u03bf\u03c3\u03cc", "Amount")
    H_NOTES   = t("\u03a3\u03b7\u03bc\u03b5\u03b9\u03ce\u03c3\u03b5\u03b9\u03c2", "Notes")

    # ── OVERVIEW (parametric — grand totals across ALL projects, no project names) ──
    ws.row_dimensions[3].height = 24
    ws.merge_cells("A3:F3")
    c = ws.cell(row=3, column=1, value="  " + t("\u0393\u0395\u039d\u0399\u039a\u0397 \u0395\u0399\u039a\u039f\u039d\u0391 \u2014 \u039f\u039b\u0391 \u03a4\u0391 \u0395\u03a1\u0393\u0391", "OVERVIEW \u2014 ALL PROJECTS"))
    c.fill = F(LIGHT_BLUE); c.font = Fnt(size=10, bold=True, color=DARK_BLUE)
    c.alignment = Aln(h="left", v="center")
    c.border = Border(top=Side(border_style="thin", color="FFCCD5DF"),
                      bottom=Side(border_style="medium", color=MED_BLUE),
                      left=Side(border_style="thin", color="FFCCD5DF"),
                      right=Side(border_style="thin", color="FFCCD5DF"))

    exp_metric(ws, 4, M_PAID,
               f'=SUMIF(Table_Exp[{H_STATUS}],"{S_PAID}",Table_Exp[{H_AMOUNT}])',
               M_PAID_N, bg=GREEN_BG, fg=GREEN_FG)
    exp_metric(ws, 5, M_UPCOM,
               f'=SUMIF(Table_Exp[{H_STATUS}],"{S_UPCOM}",Table_Exp[{H_AMOUNT}])',
               M_UP_N, bg=ORANGE_LIGHT, fg="FF7A3800")
    exp_metric(ws, 6, M_GRAND, "=D4+D5",
               t("\u03a3\u03cd\u03bd\u03bf\u03bb\u03bf \u03cc\u03bb\u03c9\u03bd \u03c4\u03c9\u03bd \u03ad\u03c1\u03b3\u03c9\u03bd (\u03c0\u03bb\u03b7\u03c1\u03c9\u03bc\u03ad\u03bd\u03b1 + \u03b5\u03ba\u03ba\u03c1\u03b5\u03bc\u03ae)", "All projects (paid + upcoming)"),
               bg=LIGHT_BLUE, fg=DARK_BLUE, height=28)

    for rr in range(7, 11):
        ws.row_dimensions[rr].height = 8
        for ci in range(1, 7): ws.cell(row=rr, column=ci).fill = F(WHITE)

    ws.row_dimensions[11].height = 10
    for ci in range(1, 7): ws.cell(row=11, column=ci).fill = F(WHITE)

    ws.row_dimensions[12].height = 22
    headers = [H_PROJECT, H_ITEM, H_STATUS, H_SOURCE, H_AMOUNT, H_NOTES]
    for ci, txt in enumerate(headers, 1):
        c = ws.cell(row=12, column=ci, value=txt)
        c.fill = F(DARK_BLUE); c.font = Fnt(size=10, bold=True, color=WHITE)
        c.alignment = Aln(h="center", v="center"); c.border = Brd(color="FF2E75B6")

    # localized expense items
    IT_RENT  = t("\u03a0\u03c1\u03bf\u03ba\u03b1\u03c4\u03b1\u03b2\u03bf\u03bb\u03ae \u0395\u03bd\u03bf\u03b9\u03ba\u03af\u03bf\u03c5", "Rent Advance")
    IT_DEMO  = t("\u039a\u03b1\u03c4\u03b5\u03b4\u03ac\u03c6\u03b9\u03c3\u03b7 (\u0393\u03ba\u03c1\u03b5\u03bc\u03af\u03c3\u03bc\u03b1\u03c4\u03b1)", "Demolition (Gkremismata)")
    IT_HEAT  = t("\u0398\u03ad\u03c1\u03bc\u03b1\u03bd\u03c3\u03b7 / \u039c\u03c0\u03ac\u03bd\u03b9\u03bf", "Heating / Bathroom")
    IT_ENG   = t("\u039c\u03b7\u03c7\u03b1\u03bd\u03b9\u03ba\u03cc\u03c2 / \u03a0\u03b9\u03c3\u03c4\u03bf\u03c0\u03bf\u03b9\u03b7\u03c4\u03b9\u03ba\u03ac", "Engineer / Certificates")
    IT_EOT   = t("\u0395\u039f\u03a4 \u039f\u03c1\u03ad\u03c3\u03c4\u03b7\u03c2", "EOT Orestis")
    IT_ARCH  = t("\u0391\u03c1\u03c7\u03b9\u03c4\u03ad\u03ba\u03c4\u03bf\u03bd\u03b5\u03c2", "Architects")
    IT_ACC   = t("\u039b\u03bf\u03b3\u03b9\u03c3\u03c4\u03ae\u03c2 (\u0386\u03ba\u03b7\u03c2)", "Accountant (Akis)")
    IT_COMM  = t("\u039a\u03bf\u03b9\u03bd\u03cc\u03c7\u03c1\u03b7\u03c3\u03c4\u03b1", "Common Expenses")
    IT_OTAX  = t("\u03a6\u03cc\u03c1\u03bf\u03c2 \u0399\u03b4\u03b9\u03bf\u03ba\u03c4\u03ae\u03c4\u03b7 (\u0395\u03c6\u03bf\u03c1\u03af\u03b1)", "Owner Tax (Eforia)")
    IT_NOT1  = t("\u03a3\u03c5\u03bc\u03b2\u03bf\u03bb\u03b1\u03b9\u03bf\u03b3\u03c1\u03ac\u03c6\u03bf\u03c2 + \u0394\u03b9\u03ba\u03b7\u03b3\u03cc\u03c1\u03bf\u03c2", "Notary + Lawyer")
    IT_CAP   = t("\u039a\u03b5\u03c6\u03ac\u03bb\u03b1\u03b9\u03bf \u03b3\u03b9\u03b1 \u0394\u03ac\u03bd\u03b5\u03b9\u03bf", "Capital for Loan")
    IT_VAT   = t("\u03a6\u03a0\u0391 \u039a\u03b1\u03c4\u03b1\u03c3\u03ba\u03b5\u03c5\u03ae\u03c2", "Construction VAT (FPA)")
    IT_4TH   = t("\u03a0\u03c1\u03bf\u03c3\u03c6\u03bf\u03c1\u03ac 4\u03bf\u03c5 \u039f\u03c1\u03cc\u03c6\u03bf\u03c5", "4th Floor Offer")
    IT_TADV  = t("\u03a0\u03c1\u03bf\u03ba\u03b1\u03c4\u03b1\u03b2\u03bf\u03bb\u03ae \u03a6\u03cc\u03c1\u03bf\u03c5 (\u0395\u03c6\u03bf\u03c1\u03af\u03b1)", "Tax Advance (Eforia)")
    IT_NEKT  = t("\u039d\u03b5\u03ba\u03c4\u03b1\u03c1\u03af\u03b1", "Nektaria")
    IT_NOT2  = t("\u03a3\u03c5\u03bc\u03b2\u03bf\u03bb\u03b1\u03b9\u03bf\u03b3\u03c1\u03ac\u03c6\u03bf\u03c2 / \u0394\u03b9\u03ba\u03b7\u03b3\u03cc\u03c1\u03bf\u03c2", "Notary / Lawyer")
    IT_CONS  = t("\u039a\u03b1\u03c4\u03b1\u03c3\u03ba\u03b5\u03c5\u03ae", "Construction")

    N_TBD    = t("+ \u03b5\u03ba\u03ba\u03c1\u03b5\u03bc\u03bf\u03cd\u03bd \u03c0\u03bf\u03c3\u03ac \u03c0\u03c1\u03bf\u03c2 \u03ba\u03b1\u03b8\u03bf\u03c1\u03b9\u03c3\u03bc\u03cc", "+ TBD amounts pending")
    N_DONE   = t("\u039f\u03bb\u03bf\u03ba\u03bb\u03b7\u03c1\u03ce\u03b8\u03b7\u03ba\u03b5", "Complete")
    N_EOT    = t("\u20ac1.000 \u03c0\u03c1\u03bf\u03ba\u03b1\u03c4\u03b1\u03b2\u03bf\u03bb\u03ae + \u20ac1.000 \u03c5\u03c0\u03cc\u03bb\u03bf\u03b9\u03c0\u03bf", "\u20ac1,000 advance + \u20ac1,000 balance")
    N_CAP    = t("\u0391\u03c0\u03b1\u03b9\u03c4\u03b5\u03af\u03c4\u03b1\u03b9 \u03b3\u03b9\u03b1 \u03b5\u03ba\u03c4\u03b1\u03bc\u03af\u03b5\u03c5\u03c3\u03b7 \u03b4\u03b1\u03bd\u03b5\u03af\u03bf\u03c5 TEPIX", "Required to release TEPIX loan")
    N_EST    = t("\u0395\u03ba\u03c4\u03af\u03bc\u03b7\u03c3\u03b7", "Estimated")

    LAZ = "Lazaraki 5"
    AGK = "Agiou Konstantinou"
    all_data = [
        (13, LAZ, IT_RENT,  S_PAID,  SRC_PIR,  5000,   ""),
        (14, LAZ, IT_RENT,  S_PAID,  SRC_CASH, 1000,   ""),
        (15, LAZ, IT_DEMO,  S_PAID,  SRC_CASH, 2400,   N_TBD),
        (16, LAZ, IT_HEAT,  S_PAID,  SRC_CASH, 1020,   N_TBD),
        (17, LAZ, IT_ENG,   S_PAID,  SRC_CASH, 350,    N_DONE),
        (18, LAZ, IT_EOT,   S_PAID,  SRC_CASH, 2000,   N_EOT),
        (19, LAZ, IT_ARCH,  S_PAID,  SRC_CASH, 4000,   ""),
        (20, LAZ, IT_ACC,   S_PAID,  SRC_CASH, 500,    ""),
        (21, LAZ, IT_COMM,  S_PAID,  SRC_CASH, 290,    ""),
        (22, LAZ, IT_OTAX,  S_UPCOM, SRC_PIR,  39000,  ""),
        (23, LAZ, IT_NOT1,  S_UPCOM, SRC_PIR,  20000,  ""),
        (24, LAZ, IT_NOT1,  S_UPCOM, SRC_CASH, 10000,  ""),
        (25, LAZ, IT_CAP,   S_UPCOM, SRC_TBD,  350000, N_CAP),
        (26, LAZ, IT_VAT,   S_UPCOM, SRC_TBD,  290000, N_EST),
        (27, LAZ, IT_4TH,   S_UPCOM, SRC_TBD,  50000,  ""),
        (28, AGK, IT_TADV,  S_UPCOM, SRC_PIR,  17300,  ""),
        (29, AGK, IT_TADV,  S_UPCOM, SRC_PIR,  1450,   ""),
        (30, AGK, IT_NEKT,  S_UPCOM, SRC_CASH, 50000,  ""),
        (31, AGK, IT_NOT2,  S_UPCOM, SRC_CASH, 5000,   ""),
        (32, AGK, IT_CONS,  S_UPCOM, SRC_TBD,  80000,  ""),
    ]
    for row, project, item, status, source, amt, notes in all_data:
        exp_data_row(ws, row, project, item, status, source, amt, notes)

    exp_empty_rows(ws, 33, 62)

    tbl_exp = Table(displayName="Table_Exp", ref="A12:F62")
    tbl_exp.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False)
    ws.add_table(tbl_exp)

    ws.row_dimensions[63].height = 22
    ws.merge_cells("A63:F63")
    c = ws.cell(row=63, column=1,
                value="  " + t("\u03a3\u03c5\u03bc\u03b2\u03bf\u03c5\u03bb\u03ae: \u0395\u03c0\u03b9\u03bb\u03ad\u03be\u03c4\u03b5 \u03c4\u03b1 \u03ba\u03af\u03c4\u03c1\u03b9\u03bd\u03b1 \u03ba\u03b5\u03bb\u03b9\u03ac \u03b3\u03b9\u03b1 \u03bb\u03af\u03c3\u03c4\u03b5\u03c2.  \u03a0\u03b1\u03c4\u03ae\u03c3\u03c4\u03b5 Tab \u03b1\u03c0\u03cc \u03c4\u03b7\u03bd \u03c4\u03b5\u03bb\u03b5\u03c5\u03c4\u03b1\u03af\u03b1 \u03b3\u03c1\u03b1\u03bc\u03bc\u03ae \u03b3\u03b9\u03b1 \u03b1\u03c5\u03c4\u03cc\u03bc\u03b1\u03c4\u03b7 \u03c0\u03c1\u03bf\u03c3\u03b8\u03ae\u03ba\u03b7 \u03bd\u03ad\u03b1\u03c2 \u03b3\u03c1\u03b1\u03bc\u03bc\u03ae\u03c2.",
                          "Tip: Select yellow cells for dropdowns.  Press Tab from the last row to add a new row automatically."))
    c.font = Fnt(size=10, italic=True, color="FF6B7280")
    c.fill = F("FFF6F8FA"); c.alignment = Aln(h="left")

    # ── PROJECT REGISTRY (master list) — drives the dropdown, the per-project
    #    totals and the Control Center "Other projects" section.
    #    Rows 3-14 = 12 project slots (2 pre-filled + 10 spare). Type names in H. ──
    ws.column_dimensions["G"].width = 3
    ws.column_dimensions["H"].width = 24
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 14

    ws.merge_cells("H1:K1")
    rc = ws.cell(row=1, column=8,
                 value="  " + t("\u0395\u03a1\u0393\u0391 \u2014 \u03c0\u03c1\u03cc\u03c3\u03b8\u03b5\u03c3\u03b5 \u03ad\u03c9\u03c2 12 (\u03b3\u03c1\u03ac\u03c8\u03b5 \u03c3\u03c4\u03b7 \u03c3\u03c4\u03ae\u03bb\u03b7 H)",
                              "PROJECTS \u2014 add up to 12 (type in column H)"))
    rc.fill = F(DARK_BLUE); rc.font = Fnt(size=11, bold=True, color=WHITE)
    rc.alignment = Aln(h="left", v="center")

    ws.row_dimensions[2].height = 20
    for j, htxt in enumerate([t("\u0388\u03c1\u03b3\u03bf", "Project"),
                              t("\u03a0\u03bb\u03b7\u03c1\u03c9\u03bc\u03ad\u03bd\u03b1", "Paid"),
                              t("\u0395\u03ba\u03ba\u03c1\u03b5\u03bc\u03ae", "Upcoming"),
                              t("\u03a3\u03cd\u03bd\u03bf\u03bb\u03bf", "Total")]):
        hc = ws.cell(row=2, column=8 + j, value=htxt)
        hc.fill = F(LIGHT_BLUE); hc.font = Fnt(size=10, bold=True, color=DARK_GREY)
        hc.alignment = Aln(h="center"); hc.border = Brd()

    reg_base = ["Lazaraki 5", "Agiou Konstantinou"]
    for i in range(12):
        r = 3 + i
        ws.row_dimensions[r].height = 18
        nm = ws.cell(row=r, column=8, value=(reg_base[i] if i < len(reg_base) else None))
        nm.fill = F(YELLOW); nm.border = Brd(); nm.alignment = Aln(h="left")
        nm.font = Fnt(size=10, bold=(i < len(reg_base)))
        pd = ws.cell(row=r, column=9,
            value=f'=IF($H{r}="","",SUMIFS(Table_Exp[{H_AMOUNT}],Table_Exp[{H_PROJECT}],$H{r},Table_Exp[{H_STATUS}],"{S_PAID}"))')
        up = ws.cell(row=r, column=10,
            value=f'=IF($H{r}="","",SUMIFS(Table_Exp[{H_AMOUNT}],Table_Exp[{H_PROJECT}],$H{r},Table_Exp[{H_STATUS}],"{S_UPCOM}"))')
        tt = ws.cell(row=r, column=11, value=f'=IF($H{r}="","",I{r}+J{r})')
        for cc in (pd, up, tt):
            cc.number_format = EUR; cc.alignment = Aln(h="right"); cc.border = Brd()
            cc.fill = F(LIGHT_GREY); cc.font = Fnt(size=10)

    ws.row_dimensions[15].height = 20
    tl = ws.cell(row=15, column=8, value=t("\u03a3\u03a5\u039d\u039f\u039b\u039f \u039f\u039b\u03a9\u039d", "TOTAL (all)"))
    tl.fill = F(LIGHT_BLUE); tl.font = Fnt(size=10, bold=True); tl.alignment = Aln(h="left"); tl.border = Brd()
    for col, letter in [(9, "I"), (10, "J"), (11, "K")]:
        sc = ws.cell(row=15, column=col, value=f'=SUM({letter}3:{letter}14)')
        sc.number_format = EUR; sc.font = Fnt(size=10, bold=True, color=DARK_BLUE)
        sc.alignment = Aln(h="right"); sc.border = Brd(); sc.fill = F(LIGHT_BLUE)

    # Project dropdown reads the registry, so new projects appear automatically
    dv_project = DataValidation(type="list", formula1="$H$3:$H$14",
                                allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_project); dv_project.sqref = "A13:A62"

    dv_status = DataValidation(type="list", formula1=f'"{S_PAID},{S_UPCOM},{S_HOLD}"',
                               allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_status); dv_status.sqref = "C13:C62"

    dv_source = DataValidation(type="list",
                               formula1=f'"Optima Bank,{SRC_PIR},{SRC_CASH},{SRC_GOLD},{SRC_TBD}"',
                               allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_source); dv_source.sqref = "D13:D62"

    # ══════════════════════════════════════════════════════════════════════════
    # TAB: ACCOUNTS
    # ══════════════════════════════════════════════════════════════════════════
    ws = ws_acc
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 28
    ws.freeze_panes = "A3"

    ws.row_dimensions[1].height = 34
    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1, value="  " + t("\u039b\u039f\u0393\u0391\u03a1\u0399\u0391\u03a3\u039c\u039f\u0399 & \u03a5\u03a0\u039f\u039b\u039f\u0399\u03a0\u0391", "ACCOUNTS & BALANCES"))
    c.fill = F(DARK_BLUE); c.font = Fnt(size=14, bold=True, color=WHITE)
    c.alignment = Aln(h="left", v="center")

    ws.row_dimensions[2].height = 8
    for ci in range(1, 5): ws.cell(row=2, column=ci).fill = F(WHITE)

    ws.row_dimensions[3].height = 24
    ws.merge_cells("A3:D3")
    c = ws.cell(row=3, column=1, value="  " + t("\u03a4\u0391 \u03a7\u03a1\u0397\u039c\u0391\u03a4\u0391 \u03a3\u0391\u03a3 \u03a3\u0397\u039c\u0395\u03a1\u0391", "YOUR MONEY TODAY"))
    c.fill = F(LIGHT_GREY); c.font = Fnt(size=10, bold=True, color=DARK_BLUE)
    c.alignment = Aln(h="left", v="center")
    c.border = Border(top=Side(border_style="thin", color="FFCCD5DF"),
                      bottom=Side(border_style="medium", color=MED_BLUE),
                      left=Side(border_style="thin", color="FFCCD5DF"),
                      right=Side(border_style="thin", color="FFCCD5DF"))

    acc_headers = [t("\u039b\u03bf\u03b3\u03b1\u03c1\u03b9\u03b1\u03c3\u03bc\u03cc\u03c2", "Account"),
                   t("\u03a5\u03c0\u03cc\u03bb\u03bf\u03b9\u03c0\u03bf (\u20ac)", "Balance (\u20ac)"),
                   t("\u03a4\u03b5\u03bb\u03b5\u03c5\u03c4\u03b1\u03af\u03b1 \u0395\u03bd\u03b7\u03bc\u03ad\u03c1\u03c9\u03c3\u03b7", "Last Updated"),
                   t("\u03a3\u03b7\u03bc\u03b5\u03b9\u03ce\u03c3\u03b5\u03b9\u03c2", "Notes")]
    for ci, h in enumerate(acc_headers, 1):
        col_hdr(ws, 4, ci, h)

    accounts = [
        (5, "Optima Bank",                                                   135000, ""),
        (6, t("\u03a4\u03c1\u03ac\u03c0\u03b5\u03b6\u03b1 \u03a0\u03b5\u03b9\u03c1\u03b1\u03b9\u03ce\u03c2", "Piraeus Bank"), 30000, ""),
        (7, t("\u039c\u03b5\u03c4\u03c1\u03b7\u03c4\u03ac (\u03c6\u03c5\u03c3\u03b9\u03ba\u03ac)", "Cash (physical)"),   220000, t("\u039a\u03c1\u03b1\u03c4\u03ae\u03c3\u03c4\u03b5 \u03b5\u03bd\u03b7\u03bc\u03b5\u03c1\u03c9\u03bc\u03ad\u03bd\u03bf", "Keep updated")),
        (8, t("\u03a7\u03c1\u03c5\u03c3\u03cc\u03c2 (\u03b5\u03ba\u03c4. \u03b1\u03be\u03af\u03b1)", "Gold (est. value)"), 250000, t("\u0395\u03ba\u03c4\u03b9\u03bc\u03ce\u03bc\u03b5\u03bd\u03b7 \u03b1\u03b3\u03bf\u03c1\u03b1\u03af\u03b1 \u03b1\u03be\u03af\u03b1", "Estimated market value")),
    ]
    for row, account, amount, notes in accounts:
        ws.row_dimensions[row].height = 22
        lbl(ws, row, 1, account, bg=WHITE)
        inp(ws, row, 2, amount, EUR)
        lbl(ws, row, 3, "", bg=WHITE)
        lbl(ws, row, 4, notes, bg=WHITE)

    ws.row_dimensions[9].height = 8
    for ci in range(1, 5): ws.cell(row=9, column=ci).fill = F(WHITE)

    ws.row_dimensions[10].height = 30
    lbl(ws, 10, 1, t("\u03a3\u03a5\u039d\u039f\u039b\u039f \u03a3\u0397\u039c\u0395\u03a1\u0391 (\u0395\u03c0\u03b9\u03b2\u03b5\u03b2\u03b1\u03b9\u03c9\u03bc\u03ad\u03bd\u03b1)", "TOTAL TODAY (Confirmed)"), bold=True, bg=LIGHT_BLUE)
    calc(ws, 10, 2, "=SUM(B5:B8)+SUM(B23:B32)", EUR, bg=LIGHT_BLUE, bold=True, size=14, fg=DARK_BLUE)
    for ci in [3, 4]:
        ws.cell(row=10, column=ci).fill = F(LIGHT_BLUE)
        ws.cell(row=10, column=ci).border = Brd()

    ws.row_dimensions[11].height = 16
    ws.row_dimensions[12].height = 8
    for r in [11, 12]:
        for ci in range(1, 5): ws.cell(row=r, column=ci).fill = F(WHITE)

    ws.row_dimensions[13].height = 24
    ws.merge_cells("A13:D13")
    c = ws.cell(row=13, column=1, value="  " + t("\u0395\u0399\u03a3\u0395\u03a1\u03a7\u039f\u039c\u0395\u039d\u0391 \u039a\u0395\u03a6\u0391\u039b\u0391\u0399\u0391  (\u03b4\u03b5\u03bd \u03ad\u03c7\u03bf\u03c5\u03bd \u03b5\u03b9\u03c3\u03c0\u03c1\u03b1\u03c7\u03b8\u03b5\u03af)", "INCOMING FUNDS  (not yet received)"))
    c.fill = F(LIGHT_GREY); c.font = Fnt(size=10, bold=True, color=DARK_BLUE)
    c.alignment = Aln(h="left", v="center")
    c.border = Border(top=Side(border_style="thin", color="FFCCD5DF"),
                      bottom=Side(border_style="medium", color=MED_BLUE),
                      left=Side(border_style="thin", color="FFCCD5DF"),
                      right=Side(border_style="thin", color="FFCCD5DF"))

    ws.row_dimensions[14].height = 28
    lbl(ws, 14, 1, t("\u0394\u03ac\u03bd\u03b5\u03b9\u03bf TEPIX  (\u039c\u039f\u039d\u039f LAZARAKI \u2014 \u03a0\u03b5\u03b9\u03c1\u03b1\u03b9\u03ce\u03c2)", "TEPIX Loan  (LAZARAKI ONLY \u2014 Piraeus)"), bold=True, bg="FFFEF0D8")
    inp(ws, 14, 2, 850000, EUR)
    c = ws.cell(row=14, column=3, value=A_PEND)
    c.fill = F(YELLOW); c.alignment = Aln(h="center"); c.border = Brd()
    lbl(ws, 14, 4, t("\u0391\u03c0\u03bf\u03ba\u03bb\u03b5\u03b9\u03c3\u03c4\u03b9\u03ba\u03ac \u03b3\u03b9\u03b1 \u03c4\u03b7\u03bd \u03b1\u03bd\u03b1\u03ba\u03b1\u03af\u03bd\u03b9\u03c3\u03b7 Lazaraki", "Dedicated to Lazaraki renovation only"), bg="FFFEF0D8")

    ws.row_dimensions[15].height = 22
    lbl(ws, 15, 1, t("\u0391\u03bd\u03b1\u03bc\u03b5\u03bd\u03cc\u03bc\u03b5\u03bd\u03b1 (\u03ac\u03bb\u03bb\u03b1 \u03b5\u03b9\u03c3\u03b5\u03c1\u03c7\u03cc\u03bc\u03b5\u03bd\u03b1)", "Expecting (other incoming)"), bg=WHITE)
    inp(ws, 15, 2, 260000, EUR)
    c = ws.cell(row=15, column=3, value=A_PEND)
    c.fill = F(YELLOW); c.alignment = Aln(h="center"); c.border = Brd()
    lbl(ws, 15, 4, "", bg=WHITE)

    ws.row_dimensions[16].height = 8
    for ci in range(1, 5): ws.cell(row=16, column=ci).fill = F(WHITE)

    ws.row_dimensions[17].height = 30
    lbl(ws, 17, 1, t("\u03a3\u03a5\u039d\u039f\u039b\u039f \u0395\u0399\u03a3\u0395\u03a1\u03a7\u039f\u039c\u0395\u039d\u03a9\u039d (\u03b1\u03bd \u03b5\u03b9\u03c3\u03c0\u03c1\u03b1\u03c7\u03b8\u03bf\u03cd\u03bd \u03cc\u03bb\u03b1)", "TOTAL INCOMING (if all received)"), bold=True, bg=LIGHT_BLUE)
    calc(ws, 17, 2, "=SUM(B14:B15)+SUM(B36:B40)", EUR, bg=LIGHT_BLUE, bold=True, size=14, fg=DARK_BLUE)
    for ci in [3, 4]:
        ws.cell(row=17, column=ci).fill = F(LIGHT_BLUE)
        ws.cell(row=17, column=ci).border = Brd()

    ws.row_dimensions[18].height = 12
    for ci in range(1, 5): ws.cell(row=18, column=ci).fill = F(WHITE)

    ws.merge_cells("A19:D19")
    c = ws.cell(row=19, column=1,
                value="  " + t("\u03a3\u03c5\u03bc\u03b2\u03bf\u03c5\u03bb\u03ae: \u0395\u03bd\u03b7\u03bc\u03b5\u03c1\u03ce\u03bd\u03b5\u03c4\u03b5 \u03c4\u03b1 \u03ba\u03af\u03c4\u03c1\u03b9\u03bd\u03b1 \u03ba\u03b5\u03bb\u03b9\u03ac \u03cc\u03c4\u03b1\u03bd \u03b1\u03bb\u03bb\u03ac\u03b6\u03bf\u03c5\u03bd \u03c4\u03b1 \u03c5\u03c0\u03cc\u03bb\u03bf\u03b9\u03c0\u03b1.  \u0391\u03bb\u03bb\u03ac\u03be\u03c4\u03b5 \u03c4\u03b7\u03bd \u03ba\u03b1\u03c4\u03ac\u03c3\u03c4\u03b1\u03c3\u03b7 TEPIX \u03c3\u03b5 '\u0395\u03bb\u03ae\u03c6\u03b8\u03b7' \u03cc\u03c4\u03b1\u03bd \u03c6\u03c4\u03ac\u03c3\u03bf\u03c5\u03bd \u03c4\u03b1 \u03c7\u03c1\u03ae\u03bc\u03b1\u03c4\u03b1.",
                          "Tip: Update yellow cells whenever balances change.  Change TEPIX status to Received when funds arrive."))
    c.font = Fnt(size=10, italic=True, color="FF6B7280")
    c.fill = F("FFF6F8FA"); c.alignment = Aln(h="left")
    ws.row_dimensions[19].height = 20

    # ── EXTRA ACCOUNTS & INCOMING — add more here; the totals above update. ──
    def _acc_section(row, title):
        ws.row_dimensions[row].height = 24
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        hc = ws.cell(row=row, column=1, value="  " + title)
        hc.fill = F(LIGHT_GREY); hc.font = Fnt(size=10, bold=True, color=DARK_BLUE)
        hc.alignment = Aln(h="left", v="center")
        hc.border = Border(top=Side(border_style="thin", color="FFCCD5DF"),
                           bottom=Side(border_style="medium", color=MED_BLUE),
                           left=Side(border_style="thin", color="FFCCD5DF"),
                           right=Side(border_style="thin", color="FFCCD5DF"))

    ws.row_dimensions[20].height = 14
    for ci in range(1, 5): ws.cell(row=20, column=ci).fill = F(WHITE)

    _acc_section(21, t("ΕΞΤΡΑ ΛΟΓΑΡΙΑΣΜΟΙ (πρόσθεσε εδώ)", "EXTRA ACCOUNTS (add here)"))
    for ci, h in enumerate(acc_headers, 1):
        col_hdr(ws, 22, ci, h)
    for r in range(23, 33):   # 10 spare accounts -> balances B23:B32
        ws.row_dimensions[r].height = 20
        inp(ws, r, 1, None, align_h="left")
        inp(ws, r, 2, None, EUR)
        lbl(ws, r, 3, "", bg=WHITE)
        lbl(ws, r, 4, "", bg=WHITE)

    ws.row_dimensions[33].height = 14
    for ci in range(1, 5): ws.cell(row=33, column=ci).fill = F(WHITE)

    _acc_section(34, t("ΕΞΤΡΑ ΕΙΣΕΡΧΟΜΕΝΑ (πρόσθεσε εδώ)", "EXTRA INCOMING (add here)"))
    for ci, h in enumerate(acc_headers, 1):
        col_hdr(ws, 35, ci, h)
    for r in range(36, 41):   # 5 spare incoming -> amounts B36:B40
        ws.row_dimensions[r].height = 20
        inp(ws, r, 1, None, align_h="left")
        inp(ws, r, 2, None, EUR)
        st = ws.cell(row=r, column=3); st.fill = F(YELLOW)
        st.alignment = Aln(h="center"); st.border = Brd()
        lbl(ws, r, 4, "", bg=WHITE)

    dv_acc_status = DataValidation(type="list", formula1=f'"{A_PEND},{A_RECV},{A_CANC}"',
                                   allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_acc_status); dv_acc_status.sqref = "C14:C15 C36:C40"

    # ══════════════════════════════════════════════════════════════════════════
    # TAB: INFO
    # ══════════════════════════════════════════════════════════════════════════
    info_title = t("\u03a0\u03a9\u03a3 \u039d\u0391 \u03a7\u03a1\u0397\u03a3\u0399\u039c\u039f\u03a0\u039f\u0399\u0397\u03a3\u0395\u03a4\u0395 \u0391\u03a5\u03a4\u039f \u03a4\u039f \u0391\u03a1\u03a7\u0395\u0399\u039f", "HOW TO USE THIS WORKBOOK")

    _info_tab(wb, info_title, "Info", [

        (t("\u0395\u03a0\u0399\u03a3\u039a\u039f\u03a0\u0397\u03a3\u0397 \u039a\u0391\u03a1\u03a4\u0395\u039b\u03a9\u039d", "TAB OVERVIEW"), "FF1868A8", [
            ("h", t("\u03a4\u03b9 \u03ba\u03ac\u03bd\u03b5\u03b9 \u03b7 \u03ba\u03ac\u03b8\u03b5 \u03ba\u03b1\u03c1\u03c4\u03ad\u03bb\u03b1:", "What each tab does:"), ""),
            ("b", t("Control Center  \u2014  \u0397 \u03ba\u03b5\u03bd\u03c4\u03c1\u03b9\u03ba\u03ae \u03bf\u03b8\u03cc\u03bd\u03b7. \u0394\u03b5\u03af\u03c7\u03bd\u03b5\u03b9 \u03cc\u03bb\u03b1 \u03c4\u03b1 \u03c7\u03c1\u03ae\u03bc\u03b1\u03c4\u03b1 \u03ba\u03b1\u03b9 \u03c4\u03b1 \u03ad\u03be\u03bf\u03b4\u03b1 \u03bc\u03b5 \u03bc\u03b9\u03b1 \u03bc\u03b1\u03c4\u03b9\u03ac. \u039c\u03cc\u03bd\u03bf \u03b1\u03bd\u03ac\u03b3\u03bd\u03c9\u03c3\u03b7 \u2014 \u03bc\u03b7\u03bd \u03b3\u03c1\u03ac\u03c6\u03b5\u03c4\u03b5 \u03b5\u03b4\u03ce.",
                    "Control Center  \u2014  Your main dashboard. Shows all money and project costs at a glance. Read-only \u2014 do not type here."), ""),
            ("b", t("Expenses  \u2014  \u0397 \u03bb\u03af\u03c3\u03c4\u03b1 \u03cc\u03bb\u03c9\u03bd \u03c4\u03c9\u03bd \u03b4\u03b1\u03c0\u03b1\u03bd\u03ce\u03bd. \u0395\u03b4\u03ce \u03c0\u03c1\u03bf\u03c3\u03b8\u03ad\u03c4\u03b5\u03c4\u03b5, \u03b5\u03c0\u03b5\u03be\u03b5\u03c1\u03b3\u03b1\u03b6\u03cc\u03c3\u03b1\u03c3\u03c4\u03b5 \u03ba\u03b1\u03b9 \u03c0\u03b1\u03c1\u03b1\u03ba\u03bf\u03bb\u03bf\u03c5\u03b8\u03b5\u03af\u03c4\u03b5 \u03ba\u03ac\u03b8\u03b5 \u03ba\u03cc\u03c3\u03c4\u03bf\u03c2.",
                    "Expenses  \u2014  The full list of all project expenses. This is where you add, edit, and track every cost."), ""),
            ("b", t("Accounts  \u2014  \u03a4\u03c1\u03ad\u03c7\u03bf\u03bd\u03c4\u03b1 \u03c5\u03c0\u03cc\u03bb\u03bf\u03b9\u03c0\u03b1: \u03c4\u03c1\u03ac\u03c0\u03b5\u03b6\u03b5\u03c2, \u03bc\u03b5\u03c4\u03c1\u03b7\u03c4\u03ae, \u03c7\u03c1\u03c5\u03c3\u03cc\u03c2, \u03b5\u03b9\u03c3\u03b5\u03c1\u03c7\u03cc\u03bc\u03b5\u03bd\u03b1 (\u03b4\u03ac\u03bd\u03b5\u03b9\u03bf TEPIX \u03ba.\u03bb\u03c0.).",
                    "Accounts  \u2014  Your current balances: banks, cash, gold, and incoming funds (TEPIX loan, etc.)."), ""),
            ("b", t("Info  \u2014  \u0391\u03c5\u03c4\u03cc\u03c2 \u03bf \u03bf\u03b4\u03b7\u03b3\u03cc\u03c2 \u03c7\u03c1\u03ae\u03c3\u03b7\u03c2 \u03c3\u03c4\u03b1 \u0395\u03bb\u03bb\u03b7\u03bd\u03b9\u03ba\u03ac.",
                    "Info  \u2014  This guide, in English."), ""),
            ("sp", "6", ""),
            ("tip", t("\u03a4\u03bf Control Center \u03b5\u03bd\u03b7\u03bc\u03b5\u03c1\u03ce\u03bd\u03b5\u03c4\u03b1\u03b9 \u03b1\u03c5\u03c4\u03cc\u03bc\u03b1\u03c4\u03b1 \u03cc\u03c4\u03b1\u03bd \u03b1\u03bb\u03bb\u03ac\u03b6\u03b5\u03c4\u03b5 \u03ba\u03ac\u03c4\u03b9 \u03c3\u03c4\u03b1 Expenses \u03ae Accounts.",
                       "The Control Center updates automatically whenever you change anything in Expenses or Accounts."), ""),
        ]),

        (t("\u03a0\u03a9\u03a3 \u039d\u0391 \u03a0\u03a1\u039f\u03a3\u0398\u0395\u03a3\u0395\u03a4\u0395 \u039c\u0399\u0391 \u0394\u0391\u03a0\u0391\u039d\u0397", "HOW TO ADD AN EXPENSE"), "FF1868A8", [
            ("s", t("1.  \u03a0\u03b7\u03b3\u03b1\u03af\u03bd\u03b5\u03c4\u03b5 \u03c3\u03c4\u03b7\u03bd \u03ba\u03b1\u03c1\u03c4\u03ad\u03bb\u03b1 \u00abExpenses\u00bb.", "1.  Go to the \u201cExpenses\u201d tab."), ""),
            ("s", t("2.  \u039a\u03b1\u03c4\u03b5\u03b2\u03b1\u03af\u03bd\u03b5\u03c4\u03b5 \u03c3\u03c4\u03b7\u03bd \u03c0\u03c1\u03ce\u03c4\u03b7 \u03ba\u03b5\u03bd\u03ae \u03ba\u03af\u03c4\u03c1\u03b9\u03bd\u03b7 \u03b3\u03c1\u03b1\u03bc\u03bc\u03ae (\u03ba\u03ac\u03c4\u03c9 \u03b1\u03c0\u03cc \u03c4\u03b1 \u03c5\u03c0\u03ac\u03c1\u03c7\u03bf\u03bd\u03c4\u03b1 \u03b4\u03b5\u03b4\u03bf\u03bc\u03ad\u03bd\u03b1).",
                    "2.  Scroll down to the first empty yellow row (below the existing data)."), ""),
            ("s", t("3.  \u039a\u03bb\u03b9\u03ba\u03ac\u03c1\u03b5\u03c4\u03b5 \u03c4\u03bf \u03ba\u03af\u03c4\u03c1\u03b9\u03bd\u03bf \u03ba\u03b5\u03bb\u03af \u03c3\u03c4\u03b7 \u03c3\u03c4\u03ae\u03bb\u03b7 \u00abProject\u00bb \u03ba\u03b1\u03b9 \u03b5\u03c0\u03b9\u03bb\u03ad\u03be\u03c4\u03b5 \u03b1\u03c0\u03cc \u03c4\u03b7 \u03bb\u03af\u03c3\u03c4\u03b1.",
                    "3.  Click the yellow cell in the \u201cProject\u201d column \u2192 pick from the dropdown."), t("\u03c0.\u03c7.  Lazaraki 5", "e.g.  Lazaraki 5")),
            ("s", t("4.  \u0393\u03c1\u03ac\u03c8\u03c4\u03b5 \u03c0\u03b5\u03c1\u03b9\u03b3\u03c1\u03b1\u03c6\u03ae \u03c3\u03c4\u03b7\u03bd \u03b5\u03c0\u03cc\u03bc\u03b5\u03bd\u03b7 \u03c3\u03c4\u03ae\u03bb\u03b7.",
                    "4.  Type a description in the next column."), t("\u03c0.\u03c7.  \u03a5\u03b4\u03c1\u03b1\u03c5\u03bb\u03b9\u03ba\u03cc\u03c2 \u2014 \u03ba\u03bf\u03c5\u03b6\u03af\u03bd\u03b1", "e.g.  Plumber \u2014 kitchen")),
            ("s", t("5.  \u039a\u03bb\u03b9\u03ba\u03ac\u03c1\u03b5\u03c4\u03b5 \u03c4\u03bf \u03ba\u03b5\u03bb\u03af \u00abStatus\u00bb \u03ba\u03b1\u03b9 \u03b5\u03c0\u03b9\u03bb\u03ad\u03be\u03c4\u03b5 \u03ba\u03b1\u03c4\u03ac\u03c3\u03c4\u03b1\u03c3\u03b7.",
                    "5.  Click the \u201cStatus\u201d cell \u2192 pick from the dropdown."), S_PAID + "  " + t("\u03ae", "or") + "  " + S_UPCOM),
            ("s", t("6.  \u039a\u03bb\u03b9\u03ba\u03ac\u03c1\u03b5\u03c4\u03b5 \u03c4\u03bf \u03ba\u03b5\u03bb\u03af \u00abSource\u00bb \u03ba\u03b1\u03b9 \u03b5\u03c0\u03b9\u03bb\u03ad\u03be\u03c4\u03b5 \u03c0\u03bf\u03b9\u03bf\u03c2 \u03bb\u03bf\u03b3\u03b1\u03c1\u03b9\u03b1\u03c3\u03bc\u03cc\u03c2 \u03c0\u03bb\u03ae\u03c1\u03c9\u03c3\u03b5.",
                    "6.  Click the \u201cSource\u201d cell \u2192 pick which account paid."), f"{SRC_CASH}, Optima Bank \u2026"),
            ("s", t("7.  \u0393\u03c1\u03ac\u03c8\u03c4\u03b5 \u03c4\u03bf \u03c0\u03bf\u03c3\u03cc (\u03bc\u03cc\u03bd\u03bf \u03b1\u03c1\u03b9\u03b8\u03bc\u03cc\u03c2, \u03c7\u03c9\u03c1\u03af\u03c2 \u03c3\u03cd\u03bc\u03b2\u03bf\u03bb\u03bf \u20ac).",
                    "7.  Type the amount (numbers only, no \u20ac sign)."), ""),
            ("s", t("8.  \u03a0\u03c1\u03bf\u03b1\u03b9\u03c1\u03b5\u03c4\u03b9\u03ba\u03ac: \u03c0\u03c1\u03bf\u03c3\u03b8\u03ad\u03c3\u03c4\u03b5 \u03c3\u03b7\u03bc\u03b5\u03b9\u03ce\u03c3\u03b5\u03b9\u03c2 \u03c3\u03c4\u03b7\u03bd \u03c4\u03b5\u03bb\u03b5\u03c5\u03c4\u03b1\u03af\u03b1 \u03c3\u03c4\u03ae\u03bb\u03b7.",
                    "8.  Add any notes in the last column (optional)."), ""),
            ("sp", "4", ""),
            ("tip", t("\u03a4\u03b1 \u03c3\u03cd\u03bd\u03bf\u03bb\u03b1 \u03c3\u03c4\u03bf Control Center \u03b5\u03bd\u03b7\u03bc\u03b5\u03c1\u03ce\u03bd\u03bf\u03bd\u03c4\u03b1\u03b9 \u03b1\u03c5\u03c4\u03cc\u03bc\u03b1\u03c4\u03b1 \u03bc\u03cc\u03bb\u03b9\u03c2 \u03c0\u03b1\u03c4\u03ae\u03c3\u03b5\u03c4\u03b5 Enter.",
                       "The Control Center totals update the moment you finish typing."), ""),
        ]),

        (t("\u03a0\u03a9\u03a3 \u039d\u0391 \u03a0\u03a1\u039f\u03a3\u0398\u0395\u03a3\u0395\u03a4\u0395 \u039d\u0395\u039f \u0395\u03a1\u0393\u039f", "HOW TO ADD A NEW PROJECT"), "FF1868A8", [
            ("b", t("\u03a5\u03c0\u03ac\u03c1\u03c7\u03bf\u03c5\u03bd \u03b4\u03cd\u03bf \u03b2\u03ae\u03bc\u03b1\u03c4\u03b1 \u03cc\u03c4\u03b1\u03bd \u03c0\u03c1\u03bf\u03c3\u03c4\u03af\u03b8\u03b5\u03c4\u03b1\u03b9 \u03bd\u03ad\u03bf \u03ad\u03c1\u03b3\u03bf:",
                    "There are two things to do when a new project is added:"), ""),
            ("s", t("1.  \u03a0\u03b7\u03b3\u03b1\u03af\u03bd\u03b5\u03c4\u03b5 \u03c3\u03c4\u03b1 \u00abExpenses\u00bb \u03ba\u03b1\u03b9 \u03b3\u03c1\u03ac\u03c8\u03c4\u03b5 \u03c4\u03bf \u03cc\u03bd\u03bf\u03bc\u03b1 \u03c4\u03bf\u03c5 \u03bd\u03ad\u03bf\u03c5 \u03ad\u03c1\u03b3\u03bf\u03c5 \u03c3\u03c4\u03b7 \u03c3\u03c4\u03ae\u03bb\u03b7 \u00abProject\u00bb.",
                    "1.  Go to \u201cExpenses\u201d and type the new project name in the \u201cProject\u201d column of an empty row."), ""),
            ("s", t("2.  \u03a7\u03c1\u03b7\u03c3\u03b9\u03bc\u03bf\u03c0\u03bf\u03b9\u03b5\u03af\u03c4\u03b5 \u03b1\u03ba\u03c1\u03b9\u03b2\u03ce\u03c2 \u03c4\u03bf \u03af\u03b4\u03b9\u03bf \u03cc\u03bd\u03bf\u03bc\u03b1 \u03c3\u03b5 \u03ba\u03ac\u03b8\u03b5 \u03b3\u03c1\u03b1\u03bc\u03bc\u03ae \u03c4\u03bf\u03c5 \u03c3\u03c5\u03b3\u03ba\u03b5\u03ba\u03c1\u03b9\u03bc\u03ad\u03bd\u03bf\u03c5 \u03ad\u03c1\u03b3\u03bf\u03c5.",
                    "2.  Use the exact same project name every time you add an expense for it."), t("\u0397 \u03bf\u03c1\u03b8\u03bf\u03b3\u03c1\u03b1\u03c6\u03af\u03b1 \u03ad\u03c7\u03b5\u03b9 \u03c3\u03b7\u03bc\u03b1\u03c3\u03af\u03b1!", "Spelling must match!")),
            ("s", t("3.  \u03a4\u03b1 \u03c3\u03cd\u03bd\u03bf\u03bb\u03b1 \u03c3\u03c4\u03b7\u03bd \u03ba\u03bf\u03c1\u03c5\u03c6\u03ae \u03c4\u03b7\u03c2 \u03c3\u03b5\u03bb\u03af\u03b4\u03b1\u03c2 Expenses \u03c5\u03c0\u03bf\u03bb\u03bf\u03b3\u03af\u03b6\u03bf\u03bd\u03c4\u03b1\u03b9 \u03b1\u03c5\u03c4\u03cc\u03bc\u03b1\u03c4\u03b1.",
                    "3.  The summary rows at the top of the Expenses tab will auto-calculate for that project."), ""),
            ("sp", "4", ""),
            ("tip", t("\u039f\u03c1\u03b8\u03bf\u03b3\u03c1\u03b1\u03c6\u03af\u03b1: \u00abLazaraki 5\u00bb \u03ba\u03b1\u03b9 \u00ablazaraki 5\u00bb \u03b8\u03b5\u03c9\u03c1\u03bf\u03cd\u03bd\u03c4\u03b1\u03b9 \u03b4\u03b9\u03b1\u03c6\u03bf\u03c1\u03b5\u03c4\u03b9\u03ba\u03ac \u03ad\u03c1\u03b3\u03b1!",
                       "Consistency is key \u2014 \u201cLazaraki 5\u201d and \u201clazaraki 5\u201d are treated as different projects."), ""),
        ]),

        (t("\u03a0\u03a9\u03a3 \u039d\u0391 \u0395\u039d\u0397\u039c\u0395\u03a1\u03a9\u03a3\u0395\u03a4\u0395 \u03a5\u03a0\u039f\u039b\u039f\u0399\u03a0\u0391", "HOW TO UPDATE BALANCES"), "FF1868A8", [
            ("s", t("1.  \u03a0\u03b7\u03b3\u03b1\u03af\u03bd\u03b5\u03c4\u03b5 \u03c3\u03c4\u03b7\u03bd \u03ba\u03b1\u03c1\u03c4\u03ad\u03bb\u03b1 \u00abAccounts\u00bb.", "1.  Go to the \u201cAccounts\u201d tab."), ""),
            ("s", t("2.  \u039a\u03bb\u03b9\u03ba\u03ac\u03c1\u03b5\u03c4\u03b5 \u03c4\u03bf \u03ba\u03af\u03c4\u03c1\u03b9\u03bd\u03bf \u03ba\u03b5\u03bb\u03af \u03b4\u03af\u03c0\u03bb\u03b1 \u03b1\u03c0\u03cc \u03c4\u03bf\u03bd \u03bb\u03bf\u03b3\u03b1\u03c1\u03b9\u03b1\u03c3\u03bc\u03cc \u03c0\u03bf\u03c5 \u03b8\u03ad\u03bb\u03b5\u03c4\u03b5 \u03bd\u03b1 \u03b1\u03bb\u03bb\u03ac\u03be\u03b5\u03c4\u03b5.",
                    "2.  Click the yellow cell next to the account you want to update."), ""),
            ("s", t("3.  \u0393\u03c1\u03ac\u03c8\u03c4\u03b5 \u03c4\u03bf \u03bd\u03ad\u03bf \u03c5\u03c0\u03cc\u03bb\u03bf\u03b9\u03c0\u03bf \u03ba\u03b1\u03b9 \u03c0\u03b1\u03c4\u03ae\u03c3\u03c4\u03b5 Enter.",
                    "3.  Type the new balance and press Enter."), ""),
            ("s", t("4.  \u0393\u03b9\u03b1 TEPIX \u03ae \u00ab\u0391\u03bd\u03b1\u03bc\u03b5\u03bd\u03cc\u03bc\u03b5\u03bd\u03b1\u00bb: \u03b1\u03bb\u03bb\u03ac\u03be\u03c4\u03b5 \u03ba\u03b1\u03b9 \u03c4\u03b7\u03bd \u03ba\u03b1\u03c4\u03ac\u03c3\u03c4\u03b1\u03c3\u03b7.",
                    "4.  For TEPIX or \u201cExpecting\u201d funds: also update the Status column."), f"{A_PEND} \u2192 {A_RECV}"),
            ("sp", "4", ""),
            ("tip", t("\u0395\u03bd\u03b7\u03bc\u03b5\u03c1\u03ce\u03bd\u03b5\u03c4\u03b5 \u03c4\u03b1\u03ba\u03c4\u03b9\u03ba\u03ac \u03c4\u03b1 \u03c5\u03c0\u03cc\u03bb\u03bf\u03b9\u03c0\u03b1 \u2014 \u03c4\u03bf \u00ab\u0395\u03bb\u03b5\u03cd\u03b8\u03b5\u03c1\u03bf \u039c\u03b5\u03c4\u03c1\u03b7\u03c4\u03ae\u00bb \u03c3\u03c4\u03bf Control Center \u03b5\u03be\u03b1\u03c1\u03c4\u03ac\u03c4\u03b1\u03b9 \u03b1\u03c0\u03cc \u03b1\u03ba\u03c1\u03b9\u03b2\u03ae \u03c5\u03c0\u03cc\u03bb\u03bf\u03b9\u03c0\u03b1.",
                       "Update balances regularly \u2014 the Control Center \u201cFree Cash\u201d figure depends on accurate balances."), ""),
        ]),

        (t("\u039f\u0394\u0397\u0393\u039f\u03a3 \u03a7\u03a1\u03a9\u039c\u0391\u03a4\u03a9\u039d", "COLOUR GUIDE"), "FF1868A8", [
            ("clr", ("FFE8F8F1", "FF0B6640"), t("\u03a0\u03c1\u03ac\u03c3\u03b9\u03bd\u03bf = \u03a0\u03bb\u03b7\u03c1\u03ce\u03b8\u03b7\u03ba\u03b5 / \u0398\u03b5\u03c4\u03b9\u03ba\u03cc \u03bc\u03ad\u03b3\u03b5\u03b8\u03bf\u03c2", "Green background = Paid / positive / confirmed")),
            ("clr", ("FFFDF8ED", "FF7A3800"), t("\u039a\u03af\u03c4\u03c1\u03b9\u03bd\u03bf / \u03b5\u03bb\u03b5\u03cd\u03b8\u03b5\u03c1\u03bf\u03c5 \u03c7\u03c1\u03ce\u03bc\u03b1 = \u03b5\u03c0\u03b5\u03be\u03b5\u03c1\u03b3\u03ac\u03c3\u03b9\u03bc\u03bf \u03ba\u03b5\u03bb\u03af \u2014 \u03bc\u03c0\u03bf\u03c1\u03b5\u03af\u03c4\u03b5 \u03bd\u03b1 \u03b3\u03c1\u03ac\u03c8\u03b5\u03c4\u03b5 \u03b5\u03b4\u03ce", "Yellow / ivory = editable cell \u2014 you can type here")),
            ("clr", ("FFE3EFF8", "FF1A2F4A"), t("\u039c\u03c0\u03bb\u03b5 \u03c6\u03cc\u03bd\u03c4\u03bf = \u03c5\u03c0\u03bf\u03bb\u03bf\u03b3\u03b9\u03c3\u03bc\u03ad\u03bd\u03bf \u03ba\u03b5\u03bb\u03af \u2014 \u03bc\u03b7\u03bd \u03b1\u03bb\u03bb\u03ac\u03b6\u03b5\u03c4\u03b5", "Blue background = calculated formula \u2014 do not edit")),
            ("clr", ("FFFDF0D8", "FF965800"), t("\u03a0\u03bf\u03c1\u03c4\u03bf\u03ba\u03b1\u03bb\u03af = \u03a4\u03bc\u03ae\u03bc\u03b1 Lazaraki 5", "Orange tint = Lazaraki 5 section")),
            ("clr", ("FFF0EEFF", "FF5A2895"), t("\u039c\u03bf\u03b2 = \u03a4\u03bc\u03ae\u03bc\u03b1 Agiou Konstantinou", "Purple tint = Agiou Konstantinou section")),
            ("clr", ("FFFCEFEF", "FF991B1B"), t("\u039a\u03cc\u03ba\u03ba\u03b9\u03bd\u03bf = \u03c0\u03c1\u03bf\u03b5\u03b9\u03b4\u03bf\u03c0\u03bf\u03af\u03b7\u03c3\u03b7 \u03ae \u03ad\u03bb\u03bb\u03b5\u03b9\u03bc\u03bc\u03b1", "Red background = alert or shortfall")),
        ]),

        (t("\u0393\u03a1\u0397\u0393\u039f\u03a1\u0395\u03a3 \u03a3\u03a5\u039c\u0392\u039f\u03a5\u039b\u0395\u03a3", "QUICK TIPS"), "FF1868A8", [
            ("b", t("\u039a\u03af\u03c4\u03c1\u03b9\u03bd\u03b1 \u03ba\u03b5\u03bb\u03b9\u03ac \u2192 \u03b5\u03c0\u03b5\u03be\u03b5\u03c1\u03b3\u03ac\u03c3\u03b9\u03bc\u03b1.  \u0391\u03c3\u03c0\u03c1\u03bf/\u03bc\u03c0\u03bb\u03b5 \u03ba\u03b5\u03bb\u03b9\u03ac \u2192 \u03c4\u03cd\u03c0\u03bf\u03b9, \u03bc\u03b7\u03bd \u03b1\u03bb\u03bb\u03ac\u03b6\u03b5\u03c4\u03b5.",
                    "Yellow cells \u2192 editable.  White or blue cells \u2192 formulas, do not touch."), ""),
            ("b", t("\u0391\u03c0\u03bf\u03b8\u03b7\u03ba\u03b5\u03cd\u03b5\u03c4\u03b5 \u03c0\u03ac\u03bd\u03c4\u03b1 \u03bc\u03b5\u03c4\u03ac \u03b1\u03c0\u03cc \u03b1\u03bb\u03bb\u03b1\u03b3\u03ad\u03c2 (Ctrl + S).",
                    "Always save the file after making changes (Ctrl + S)."), ""),
            ("b", t("\u039c\u03b7\u03bd \u03c0\u03c1\u03bf\u03c3\u03b8\u03ad\u03c4\u03b5\u03c4\u03b5 \u03ae \u03b4\u03b9\u03b1\u03b3\u03c1\u03ac\u03c8\u03b5\u03c4\u03b5 \u03c3\u03c4\u03ae\u03bb\u03b5\u03c2 \u2014 \u03bf\u03b9 \u03c4\u03cd\u03c0\u03bf\u03b9 \u03b5\u03be\u03b1\u03c1\u03c4\u03ce\u03bd\u03c4\u03b1\u03b9 \u03b1\u03c0\u03cc \u03c4\u03b9\u03c2 \u03b8\u03ad\u03c3\u03b5\u03b9\u03c2 \u03c4\u03c9\u03bd \u03c3\u03c4\u03b7\u03bb\u03ce\u03bd.",
                    "Do not add or delete columns \u2014 the formulas depend on the column positions."), ""),
            ("b", t("\u0391\u03bd \u03ba\u03ac\u03c0\u03bf\u03b9\u03bf \u03c3\u03cd\u03bd\u03bf\u03bb\u03bf \u03c6\u03b1\u03af\u03bd\u03b5\u03c4\u03b1\u03b9 \u03bb\u03ac\u03b8\u03bf\u03c2, \u03b5\u03bb\u03ad\u03b3\u03be\u03c4\u03b5 \u03c4\u03b7\u03bd \u03bf\u03c1\u03b8\u03bf\u03b3\u03c1\u03b1\u03c6\u03af\u03b1 \u03c4\u03bf\u03c5 \u03bf\u03bd\u03cc\u03bc\u03b1\u03c4\u03bf\u03c2 \u03ad\u03c1\u03b3\u03bf\u03c5.",
                    "If a total looks wrong, check that the Project name spelling matches exactly."), ""),
            ("b", t("\u0391\u03bd\u03c4\u03af\u03b3\u03c1\u03b1\u03c6\u03b5 \u03c4\u03bf \u03b1\u03c1\u03c7\u03b5\u03af\u03bf \u03c4\u03b1\u03ba\u03c4\u03b9\u03ba\u03ac (USB \u03ae cloud).",
                    "Back up the file regularly (copy to a USB stick or cloud folder)."), ""),
        ]),
    ])

    return wb


# ─── SAVE — two files (Greek + English) ──────────────────────────────────────
if __name__ == "__main__":
    _base_dir = os.path.dirname(os.path.abspath(__file__))
    for _lang, _fname in [("gr", "LA_Budgeting_GR.xlsx"), ("en", "LA_Budgeting_EN.xlsx")]:
        LANG = _lang
        _wb = build_workbook()
        _path = os.path.join(_base_dir, _fname)
        _wb.save(_path)
        print(f"Saved: {_path}")
